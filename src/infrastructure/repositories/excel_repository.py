"""
Camada de Infraestrutura - Repositório Excel
"""
from typing import List

import openpyxl
from openpyxl import Workbook


class ExcelRepository:
    """Repositório para salvar dados em Excel"""
    
    def __init__(self, file_path: str = "emails.xlsx"):
        self.file_path = file_path
    
    def save_emails(self, emails: List[Email]) -> bool:
        """Salva e-mails em planilha Excel"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Emails"
            
            # Cabeçalhos
            ws['A1'] = "Remetente"
            ws['B1'] = "Assunto"
            ws['C1'] = "Domínio"
            
            # Dados
            for idx, email in enumerate(emails, start=2):
                ws[f'A{idx}'] = email.sender
                ws[f'B{idx}'] = email.subject
                ws[f'C{idx}'] = email.domain
            
            wb.save(self.file_path)
            return True
            
        except Exception:
            return False
    
    def load_emails(self) -> List[Email]:
        """Carrega e-mails da planilha Excel"""
        emails = []
        
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Se tem remetente
                    email = Email(
                        sender=row[0] or "",
                        subject=row[1] or "",
                        body="",
                        domain=row[2] or ""
                    )
                    emails.append(email)
                    
        except FileNotFoundError:
            pass
        
        return emails