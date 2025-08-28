"""
Camada de Infraestrutura - Persistência de dados
"""
import json
import os
from typing import Dict, Set

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

from ...domain.email_processor import Company


class JsonRepository:
    """Repositório JSON para controle de visitados e e-mails"""
    
    def __init__(self, visited_path: str, emails_path: str):
        self.visited_path = visited_path
        self.emails_path = emails_path
        # Garante que pasta data existe
        data_dir = os.path.dirname(visited_path)
        os.makedirs(data_dir, exist_ok=True)
    
    def load_visited_domains(self) -> Dict[str, bool]:
        """Carrega domínios já visitados"""
        return self._load_json(self.visited_path) or {}
    
    def save_visited_domains(self, visited: Dict[str, bool]):
        """Salva domínios visitados"""
        self._save_json(self.visited_path, visited)
    
    def load_seen_emails(self) -> Set[str]:
        """Carrega e-mails já coletados"""
        emails_list = self._load_json(self.emails_path) or []
        return set(emails_list)
    
    def save_seen_emails(self, emails: Set[str]):
        """Salva e-mails coletados"""
        self._save_json(self.emails_path, list(emails))
    
    def _load_json(self, path: str):
        """Carrega arquivo JSON"""
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except:
                return None
        return None
    
    def _save_json(self, path: str, data):
        """Salva arquivo JSON"""
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)


class ExcelRepository:
    """Repositório Excel para empresas"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self._ensure_output_dir()
        self._ensure_excel_exists()
    
    def save_company(self, company: Company):
        """Salva empresa no Excel"""
        if not company.emails:
            return
        
        emails_str = ";".join(company.emails) + ";"
        row_data = [company.url, emails_str]
        # Colunas comentadas para uso futuro:
        # row_data = [company.name, company.phone, emails_str, company.search_term, company.address, company.url]
        
        try:
            wb = load_workbook(self.file_path)
            ws = wb["Dados"]
            ws.append(row_data)
            
            # Formatação da linha
            for cell in ws[ws.max_row]:
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(horizontal="left")
            
            wb.save(self.file_path)
        except Exception:
            pass
    
    def _ensure_output_dir(self):
        """Garante que diretório de saída existe"""
        output_dir = os.path.dirname(self.file_path)
        if not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir)
                print(f"[INFO] Pasta criada: {output_dir}")
            except Exception as e:
                print(f"[ERRO] Falha ao criar pasta {output_dir}: {e}")
    
    def _ensure_excel_exists(self):
        """Garante que arquivo Excel existe"""
        if not os.path.exists(self.file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Dados"
            ws["A1"] = "site"
            ws["B1"] = "email"
            # Colunas comentadas para uso futuro:
            # ws["C1"] = "NOME DO SITE"
            # ws["D1"] = "TELEFONE"
            # ws["E1"] = "TERMO BUSCA"
            # ws["F1"] = "ENDEREÇO"
            
            # Ajusta largura das colunas
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 40
            # ws.column_dimensions['C'].width = 30
            # ws.column_dimensions['D'].width = 20
            # ws.column_dimensions['E'].width = 40
            # ws.column_dimensions['F'].width = 50
            
            # Formatação do cabeçalho
            for cell in ws[1]:
                cell.font = Font(name="Arial", size=10, bold=True)
                cell.alignment = Alignment(horizontal="left")
            
            wb.save(self.file_path)
    
