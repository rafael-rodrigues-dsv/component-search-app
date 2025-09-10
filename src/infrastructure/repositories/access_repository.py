"""
Repositório para acesso ao banco Access - Substitui JSON
"""
import logging
from pathlib import Path
from typing import List, Dict, Any, Tuple

import pyodbc


class AccessRepository:
    """Repositório principal para banco Access"""
    
    _connection = None  # Singleton de conexão
    _instance = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    def __init__(self):
        if hasattr(self, '_initialized'):
            return
        
        # Sempre usar pasta data do projeto (fora da virtualização)
        project_root = Path.cwd()
        self.db_path = project_root / "data" / "pythonsearch.accdb"
        self.db_path = self.db_path.resolve()
        self.conn_str = f'DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={self.db_path};'
        self.logger = logging.getLogger(__name__)
        self._initialized = True

    def _get_connection(self):
        """Obtém conexão singleton com o banco"""
        if self._connection is None:
            self._connection = pyodbc.connect(self.conn_str)
        return self._connection
    
    def close_connection(self):
        """Fecha conexão singleton"""
        if self._connection:
            self._connection.close()
            self._connection = None

    # ===== EMPRESAS =====

    def is_domain_visited(self, domain: str) -> bool:
        """Verifica se domínio já foi visitado (substitui visited.json)"""
        try:
            conn = self._get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM TB_EMPRESAS WHERE DOMINIO = ?", (domain,))
            result = cursor.fetchone()
            cursor.close()
            return result[0] > 0
        except Exception as e:
            self.logger.error(f"Erro ao verificar domínio visitado: {e}")
            return False

    def save_endereco(self, address_model) -> int:
        """Salva endereço estruturado e retorna ID"""
        if not address_model or not address_model.is_valid():
            return None
            
        try:
            conn = self._get_connection()
            cursor = conn.cursor()
            
            # Verificar se tabela existe
            try:
                cursor.execute("SELECT COUNT(*) FROM TB_ENDERECOS")
            except:
                print("[AVISO] Tabela TB_ENDERECOS não existe - pulando endereço")
                cursor.close()
                return None
            
            # Verificar se endereço já existe (por logradouro + numero + complemento)
            cursor.execute("""
                           SELECT ID_ENDERECO FROM TB_ENDERECOS 
                           WHERE LOGRADOURO = ? AND NUMERO = ? AND COMPLEMENTO = ? AND BAIRRO = ?
                           """, (address_model.logradouro, address_model.numero, address_model.complemento, address_model.bairro))
            existing = cursor.fetchone()
            
            if existing:
                cursor.close()
                return existing[0]
            
            # Inserir novo endereço
            cursor.execute("""
                           INSERT INTO TB_ENDERECOS (LOGRADOURO, NUMERO, COMPLEMENTO, BAIRRO, CIDADE, ESTADO, CEP, DATA_CRIACAO)
                           VALUES (?, ?, ?, ?, ?, ?, ?, Date())
                           """, (address_model.logradouro, address_model.numero, address_model.complemento, address_model.bairro, 
                           address_model.cidade, address_model.estado, address_model.cep))
            
            cursor.execute("SELECT @@IDENTITY")
            endereco_id = cursor.fetchone()[0]
            conn.commit()
            cursor.close()
            return endereco_id
        except Exception as e:
            print(f"[AVISO] Erro ao salvar endereço: {e} - continuando sem endereço")
            return None

    def save_empresa(self, termo_id: int, site_url: str, domain: str, motor_busca: str,
                     address_model = None, latitude: float = None, longitude: float = None,
                     distancia_km: float = None) -> int:
        """Salva empresa com endereço estruturado"""
        conn = self._get_connection()
        cursor = conn.cursor()
        
        # Salvar endereço se houver
        endereco_id = None
        if address_model:
            endereco_id = self.save_endereco(address_model)
        
        cursor.execute("""
                       INSERT INTO TB_EMPRESAS (ID_TERMO, SITE_URL, DOMINIO, STATUS_COLETA,
                                                DATA_PRIMEIRA_VISITA, TENTATIVAS_COLETA, MOTOR_BUSCA,
                                                ID_ENDERECO, LATITUDE, LONGITUDE, DISTANCIA_KM)
                       VALUES (?, ?, ?, ?, Date (), ?, ?, ?, ?, ?, ?)
                       """, (termo_id, site_url, domain, 'PENDENTE', 0, motor_busca, endereco_id, latitude, longitude, distancia_km))
        cursor.execute("SELECT @@IDENTITY")
        empresa_id = cursor.fetchone()[0]
        conn.commit()
        cursor.close()
        return empresa_id

    def update_empresa_status(self, empresa_id: int, status: str, nome_empresa: str = None):
        """Atualiza status da empresa após coleta"""
        conn = self._get_connection()
        cursor = conn.cursor()
        if nome_empresa:
            cursor.execute("""
                           UPDATE TB_EMPRESAS
                           SET STATUS_COLETA      = ?,
                               NOME_EMPRESA       = ?,
                               DATA_ULTIMA_VISITA = Date (), TENTATIVAS_COLETA = TENTATIVAS_COLETA + 1
                           WHERE ID_EMPRESA = ?
                           """, (status, nome_empresa, empresa_id))
        else:
            cursor.execute("""
                           UPDATE TB_EMPRESAS
                           SET STATUS_COLETA      = ?,
                               DATA_ULTIMA_VISITA = Date (), TENTATIVAS_COLETA = TENTATIVAS_COLETA + 1
                           WHERE ID_EMPRESA = ?
                           """, (status, empresa_id))
        conn.commit()
        cursor.close()

    # ===== E-MAILS =====

    def is_email_collected(self, email: str) -> bool:
        """Verifica se e-mail já foi coletado (substitui emails.json)"""
        try:
            conn = self._get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM TB_EMAILS WHERE EMAIL = ?", (email,))
            result = cursor.fetchone()
            cursor.close()
            return result[0] > 0
        except Exception as e:
            self.logger.error(f"Erro ao verificar e-mail coletado: {e}")
            return False

    def save_emails(self, empresa_id: int, emails: List[str], domain_email: str):
        """Salva e-mails otimizado"""
        if not emails:
            return

        conn = self._get_connection()
        cursor = conn.cursor()

        # Inserção em lote
        email_data = [(empresa_id, email, domain_email, -1, 'SCRAPING') for email in emails]
        cursor.executemany("""
                           INSERT INTO TB_EMAILS (ID_EMPRESA, EMAIL, DOMINIO_EMAIL,
                                                  VALIDADO, DATA_COLETA, ORIGEM_COLETA)
                           VALUES (?, ?, ?, ?, Date (), ?)
                           """, email_data)
        conn.commit()
        cursor.close()

    def save_telefones(self, empresa_id: int, telefones: List[Dict[str, str]]):
        """Salva telefones otimizado"""
        if not telefones:
            return

        conn = self._get_connection()
        cursor = conn.cursor()

        # Inserção em lote
        phone_data = [(empresa_id, tel['original'], tel['formatted'],
                       tel.get('ddd', ''), tel.get('tipo', 'FIXO'), -1) for tel in telefones]
        cursor.executemany("""
                           INSERT INTO TB_TELEFONES (ID_EMPRESA, TELEFONE, TELEFONE_FORMATADO,
                                                     DDD, TIPO_TELEFONE, VALIDADO, DATA_COLETA)
                           VALUES (?, ?, ?, ?, ?, ?, Date () )
                           """, phone_data)
        conn.commit()
        cursor.close()

    # ===== TERMOS DE BUSCA =====

    def get_pending_terms(self) -> List[Dict[str, Any]]:
        """Obtém termos pendentes de processamento"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                           SELECT ID_TERMO, TERMO_COMPLETO, TIPO_LOCALIZACAO, STATUS_PROCESSAMENTO
                           FROM TB_TERMOS_BUSCA
                           WHERE STATUS_PROCESSAMENTO = 'PENDENTE'
                           ORDER BY ID_TERMO
                           """)
            return [{'id': row[0], 'termo': row[1], 'tipo': row[2], 'status': row[3]}
                    for row in cursor.fetchall()]

    def update_term_status(self, termo_id: int, status: str):
        """Atualiza status do termo"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                           UPDATE TB_TERMOS_BUSCA
                           SET STATUS_PROCESSAMENTO = ?,
                               DATA_PROCESSAMENTO   = Date ()
                           WHERE ID_TERMO = ?
                           """, status, termo_id)
            conn.commit()

    def generate_search_terms(self):
        """Método legado - não faz nada (descoberta dinâmica substituiu)"""
        print("[INFO] Termos serão gerados dinamicamente durante a coleta")
        return 0

    # ===== RESET =====

    def reset_collected_data(self):
        """Reset rápido - limpa apenas dados de coleta, CEP e geolocalização"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            
            # Limpar apenas as 6 tabelas de dados coletados
            tables = [
                "TB_CEP_ENRICHMENT",
                "TB_GEOLOCALIZACAO", 
                "TB_TELEFONES",
                "TB_EMAILS",
                "TB_EMPRESAS",
                "TB_PLANILHA"
            ]
            
            for table in tables:
                cursor.execute(f"DELETE FROM {table}")
            
            # Resetar status dos termos para reprocessar
            cursor.execute("UPDATE TB_TERMOS_BUSCA SET STATUS_PROCESSAMENTO = 'PENDENTE', DATA_PROCESSAMENTO = NULL")
            
            conn.commit()
            print("[RESET] Dados de coleta limpos")
    
    def clear_search_terms(self):
        """Limpa termos de busca existentes"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM TB_TERMOS_BUSCA")
            conn.commit()
    
    def save_dynamic_search_terms(self, terms: list) -> int:
        """Salva termos de busca gerados dinamicamente"""
        if not terms:
            return 0
            
        with self._get_connection() as conn:
            cursor = conn.cursor()
            
            # Inserir termos em lote
            for term in terms:
                cursor.execute("""
                    INSERT INTO TB_TERMOS_BUSCA 
                    (TERMO_COMPLETO, TIPO_LOCALIZACAO, STATUS_PROCESSAMENTO, DATA_CRIACAO)
                    VALUES (?, ?, 'PENDENTE', Date())
                """, 
                term['termo'], 
                term['tipo_localizacao']
                )
            
            conn.commit()
            return len(terms)
    
    def execute_query(self, query: str, params: list = None):
        """Executa query genérica"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            
            # Se é SELECT, retornar resultados
            if query.strip().upper().startswith('SELECT'):
                columns = [desc[0] for desc in cursor.description]
                rows = cursor.fetchall()
                return [dict(zip(columns, row)) for row in rows]
            else:
                conn.commit()
                return cursor.rowcount

    # ===== PLANILHA =====

    def save_to_final_sheet(self, site_url: str, emails_str: str, telefones_str: str,
                            distancia_km: float = None):
        """Salva/atualiza registro na tabela planilha (endereço vem da TB_ENDERECOS)"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            
            # Buscar e concatenar endereço da empresa
            cursor.execute("""
                           SELECT e.LOGRADOURO, e.NUMERO, e.COMPLEMENTO, e.BAIRRO, e.CIDADE, e.ESTADO
                           FROM TB_EMPRESAS emp 
                           LEFT JOIN TB_ENDERECOS e ON emp.ID_ENDERECO = e.ID_ENDERECO 
                           WHERE emp.SITE_URL = ?
                           """, site_url)
            endereco_result = cursor.fetchone()
            
            if endereco_result and endereco_result[0]:  # Se tem logradouro
                logr, num, complemento, bairro, cidade, estado = endereco_result
                parts = []
                if logr:
                    if num:
                        numero_completo = num
                        if complemento:
                            numero_completo += f" {complemento}"
                        parts.append(f"{logr}, {numero_completo}")
                    else:
                        parts.append(logr)
                if bairro:
                    parts.append(bairro)
                if cidade:
                    parts.append(cidade)
                if estado:
                    parts.append(estado)
                endereco = ", ".join(parts)
                # Garantir limite de 255 caracteres
                if len(endereco) > 255:
                    endereco = endereco[:252] + "..."
            else:
                endereco = ""

            # Verificar se já existe
            cursor.execute("SELECT ID_PLANILHA FROM TB_PLANILHA WHERE SITE = ?", site_url)
            existing = cursor.fetchone()

            if existing:
                # Atualizar
                if distancia_km is not None:
                    cursor.execute("""
                                   UPDATE TB_PLANILHA
                                   SET EMAIL            = ?,
                                       TELEFONE         = ?,
                                       ENDERECO         = ?,
                                       DISTANCIA_KM     = ?,
                                       DATA_ATUALIZACAO = Date ()
                                   WHERE SITE = ?
                                   """, emails_str, telefones_str, endereco, distancia_km, site_url)
                else:
                    cursor.execute("""
                                   UPDATE TB_PLANILHA
                                   SET EMAIL            = ?,
                                       TELEFONE         = ?,
                                       ENDERECO         = ?,
                                       DATA_ATUALIZACAO = Date ()
                                   WHERE SITE = ?
                                   """, emails_str, telefones_str, endereco, site_url)
            else:
                # Inserir novo
                if distancia_km is not None:
                    cursor.execute("""
                                   INSERT INTO TB_PLANILHA (SITE, EMAIL, TELEFONE, ENDERECO, DISTANCIA_KM, DATA_ATUALIZACAO)
                                   VALUES (?, ?, ?, ?, ?, Date () )
                                   """, site_url, emails_str, telefones_str, endereco, distancia_km)
                else:
                    cursor.execute("""
                                   INSERT INTO TB_PLANILHA (SITE, EMAIL, TELEFONE, ENDERECO, DATA_ATUALIZACAO)
                                   VALUES (?, ?, ?, ?, Date () )
                                   """, site_url, emails_str, telefones_str, endereco)

            conn.commit()

    # ===== GEOLOCALIZACAO =====

    def create_geolocation_task(self, empresa_id: int, endereco_id: int):
        """Cria tarefa de geolocalização usando ID do endereço"""
        if not endereco_id:
            return
            
        with self._get_connection() as conn:
            cursor = conn.cursor()
            
            # Verificar se já existe tarefa para este endereço
            cursor.execute("SELECT ID_GEO FROM TB_GEOLOCALIZACAO WHERE ID_ENDERECO = ?", endereco_id)
            existing = cursor.fetchone()
            
            if not existing:
                cursor.execute("""
                               INSERT INTO TB_GEOLOCALIZACAO (ID_EMPRESA, ID_ENDERECO, STATUS_PROCESSAMENTO, TENTATIVAS)
                               VALUES (?, ?, 'PENDENTE', 0)
                               """, empresa_id, endereco_id)
                conn.commit()

    def get_pending_geolocation_tasks(self) -> List[Dict[str, Any]]:
        """Obtém tarefas de geolocalização pendentes com dados estruturados"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                           SELECT g.ID_GEO, g.ID_EMPRESA, g.ID_ENDERECO, emp.SITE_URL,
                                  end.LOGRADOURO, end.NUMERO, end.COMPLEMENTO, end.BAIRRO, end.CIDADE, end.ESTADO, end.CEP
                           FROM (TB_GEOLOCALIZACAO g 
                           INNER JOIN TB_EMPRESAS emp ON g.ID_EMPRESA = emp.ID_EMPRESA)
                           INNER JOIN TB_ENDERECOS end ON g.ID_ENDERECO = end.ID_ENDERECO
                           WHERE g.STATUS_PROCESSAMENTO = 'PENDENTE'
                           ORDER BY g.ID_GEO
                           """)
            
            tasks = []
            for row in cursor.fetchall():
                # Criar AddressModel a partir dos dados
                from src.domain.models.address_model import AddressModel
                address = AddressModel(
                    logradouro=row[4] or "",
                    numero=row[5] or "",
                    complemento=row[6] or "",
                    bairro=row[7] or "",
                    cidade=row[8] or "",
                    estado=row[9] or "",
                    cep=row[10] or ""
                )
                
                tasks.append({
                    'id_geo': row[0],
                    'id_empresa': row[1], 
                    'id_endereco': row[2],
                    'site_url': row[3],
                    'address_model': address
                })
            
            return tasks

    def update_geolocation_success(self, id_geo: int, latitude: float, longitude: float, distancia_km: float):
        """Atualiza resultado da geolocalização com sucesso"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            
            # Atualizar tabela de controle com status CONCLUIDO
            cursor.execute("""
                           UPDATE TB_GEOLOCALIZACAO
                           SET LATITUDE = ?, LONGITUDE = ?, DISTANCIA_KM = ?,
                               STATUS_PROCESSAMENTO = 'CONCLUIDO', DATA_PROCESSAMENTO = Date(),
                               TENTATIVAS = TENTATIVAS + 1
                           WHERE ID_GEO = ?
                           """, latitude, longitude, distancia_km, id_geo)
            
            # Obter ID da empresa
            cursor.execute("SELECT ID_EMPRESA FROM TB_GEOLOCALIZACAO WHERE ID_GEO = ?", id_geo)
            empresa_id = cursor.fetchone()[0]
            
            # Replicar para TB_EMPRESAS
            cursor.execute("""
                           UPDATE TB_EMPRESAS
                           SET LATITUDE = ?, LONGITUDE = ?, DISTANCIA_KM = ?
                           WHERE ID_EMPRESA = ?
                           """, latitude, longitude, distancia_km, empresa_id)
            
            # Replicar distância para TB_PLANILHA
            cursor.execute("SELECT SITE_URL FROM TB_EMPRESAS WHERE ID_EMPRESA = ?", empresa_id)
            site_result = cursor.fetchone()
            if site_result:
                site_url = site_result[0]
                cursor.execute("""
                               UPDATE TB_PLANILHA
                               SET DISTANCIA_KM = ?
                               WHERE SITE = ?
                               """, distancia_km, site_url)
            
            conn.commit()
    
    def update_geolocation_result(self, id_geo: int, latitude: float, longitude: float, distancia_km: float):
        """Método legado - usar update_geolocation_success"""
        self.update_geolocation_success(id_geo, latitude, longitude, distancia_km)
        print(f"      📋 Dados replicados: TB_EMPRESAS + TB_PLANILHA")

    def update_geolocation_error(self, id_geo: int, erro_descricao: str):
        """Atualiza erro na geolocalização"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                           UPDATE TB_GEOLOCALIZACAO
                           SET STATUS_PROCESSAMENTO = 'ERRO', DATA_PROCESSAMENTO = Date(),
                               TENTATIVAS = TENTATIVAS + 1, ERRO_DESCRICAO = ?
                           WHERE ID_GEO = ?
                           """, erro_descricao, id_geo)
            conn.commit()

    def update_planilha_distance_by_empresa(self, empresa_id: int, distancia_km: float):
        """Atualiza distância na planilha baseado no ID da empresa (método legado)"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            
            # Obter site_url da empresa
            cursor.execute("SELECT SITE_URL FROM TB_EMPRESAS WHERE ID_EMPRESA = ?", empresa_id)
            result = cursor.fetchone()
            if result:
                site_url = result[0]
                cursor.execute("""
                               UPDATE TB_PLANILHA
                               SET DISTANCIA_KM = ?
                               WHERE SITE = ?
                               """, distancia_km, site_url)
                conn.commit()
                print(f"      📋 TB_PLANILHA atualizada: {distancia_km}km para {site_url}")

    def get_geolocation_stats(self) -> Dict[str, int]:
        """Obtém estatísticas de geolocalização"""
        try:
            conn = self._get_connection()
            cursor = conn.cursor()
            
            # Total de empresas com endereço
            cursor.execute("SELECT COUNT(*) FROM TB_EMPRESAS WHERE ID_ENDERECO IS NOT NULL")
            total_com_endereco = cursor.fetchone()[0]
            
            # Tarefas de geolocalização
            cursor.execute("SELECT COUNT(*) FROM TB_GEOLOCALIZACAO WHERE STATUS_PROCESSAMENTO = 'CONCLUIDO'")
            geocodificadas = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM TB_GEOLOCALIZACAO WHERE STATUS_PROCESSAMENTO = 'PENDENTE'")
            pendentes = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM TB_GEOLOCALIZACAO WHERE STATUS_PROCESSAMENTO = 'ERRO'")
            erros = cursor.fetchone()[0]
            
            cursor.close()
            
            return {
                'total_com_endereco': total_com_endereco,
                'geocodificadas': geocodificadas,
                'pendentes': pendentes,
                'erros': erros,
                'percentual': round((geocodificadas / max(total_com_endereco, 1)) * 100, 1)
            }
        except Exception as e:
            print(f"[AVISO] Erro ao obter estatísticas: {e}")
            return {
                'total_com_endereco': 0,
                'geocodificadas': 0,
                'pendentes': 0,
                'erros': 0,
                'percentual': 0
            }

    # ===== EXCEL EXPORT =====

    def export_to_excel(self, excel_path: str):
        """Exporta dados para Excel diretamente da tabela planilha"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                           SELECT SITE, EMAIL, TELEFONE, ENDERECO, DISTANCIA_KM
                           FROM TB_PLANILHA
                           ORDER BY DISTANCIA_KM, SITE
                           """)

            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Empresas"

            # Cabeçalho
            ws['A1'] = 'SITE'
            ws['B1'] = 'EMAIL'
            ws['C1'] = 'TELEFONE'
            ws['D1'] = 'ENDERECO'
            ws['E1'] = 'DISTANCIA_KM'

            # Dados
            row = 2
            for site, emails, telefones, endereco, distancia in cursor.fetchall():
                ws[f'A{row}'] = site
                ws[f'B{row}'] = emails or ''
                ws[f'C{row}'] = telefones or ''
                ws[f'D{row}'] = endereco or ''
                ws[f'E{row}'] = distancia or ''
                row += 1

            wb.save(excel_path)
            return row - 2  # Número de registros
    
    # ===== GEOGRAPHIC DISCOVERY =====
    
    def save_discovered_cities(self, cities: List[Dict], uf: str) -> int:
        """Salva cidades descobertas dinamicamente"""
        saved_count = 0
        with self._get_connection() as conn:
            cursor = conn.cursor()
            
            for city in cities:
                try:
                    cursor.execute(
                        "INSERT INTO TB_CIDADES (NOME_CIDADE, UF, ATIVO, DATA_CRIACAO) VALUES (?, ?, ?, Date())",
                        [city['name'], uf, -1]
                    )
                    saved_count += 1
                except Exception:
                    # Ignora se já existe (duplicata)
                    pass
            
            conn.commit()
        return saved_count
    
    def save_discovered_neighborhoods(self, neighborhoods: List[Dict], uf: str) -> int:
        """Salva bairros descobertos dinamicamente"""
        saved_count = 0
        with self._get_connection() as conn:
            cursor = conn.cursor()
            
            for neighborhood in neighborhoods:
                try:
                    cursor.execute(
                        "INSERT INTO TB_BAIRROS (NOME_BAIRRO, UF, ATIVO, DATA_CRIACAO) VALUES (?, ?, ?, Date())",
                        [neighborhood['name'], uf, -1]
                    )
                    saved_count += 1
                except Exception:
                    # Ignora se já existe (duplicata)
                    pass
            
            conn.commit()
        return saved_count
    
    def create_cities_cache_table(self) -> None:
        """Cria tabela de cache de cidades (SQLite)"""
        import sqlite3
        cache_path = Path("data/cache/cities_brazil.db")
        cache_path.parent.mkdir(exist_ok=True)
        
        conn = sqlite3.connect(cache_path)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS cities (
                id TEXT PRIMARY KEY,
                name TEXT,
                state TEXT,
                population INTEGER,
                is_capital BOOLEAN,
                region_type TEXT
            )
        """)
        conn.commit()
        conn.close()
    
    def save_cities_to_cache(self, cities: List[Dict], uf: str) -> None:
        """Salva cidades no cache SQLite"""
        import sqlite3
        cache_path = Path("data/cache/cities_brazil.db")
        
        conn = sqlite3.connect(cache_path)
        for city in cities:
            conn.execute("""
                INSERT OR REPLACE INTO cities 
                (id, name, state, population, is_capital, region_type)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                city['id'],
                city['nome'],
                uf,
                city.get('population', 0),
                city.get('is_capital', False),
                'unknown'
            ))
        conn.commit()
        conn.close()
    
    def get_cities_from_cache(self, uf: str) -> List[Dict]:
        """Busca cidades do cache SQLite"""
        import sqlite3
        cache_path = Path("data/cache/cities_brazil.db")
        
        conn = sqlite3.connect(cache_path)
        conn.row_factory = sqlite3.Row
        
        cursor = conn.execute("""
            SELECT * FROM cities 
            WHERE state = ? 
            ORDER BY population DESC
        """, (uf,))
        
        cities = []
        for row in cursor:
            cities.append({
                'id': row['id'],
                'nome': row['name'],
                'population': row['population'],
                'is_capital': bool(row['is_capital'])
            })
        
        conn.close()
        print(f"[CACHE] ✅ {len(cities)} cidades carregadas de {uf} (cache local)")
        return cities
    
    def get_addresses_with_cep_for_enrichment(self) -> List[Tuple[int, str, str]]:
        """Busca endereços com CEP que ainda não foram geocodificados"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT e.ID_EMPRESA, en.LOGRADOURO, en.NUMERO, en.BAIRRO, en.CIDADE, en.CEP
                FROM TB_EMPRESAS e
                INNER JOIN TB_ENDERECOS en ON e.ID_ENDERECO = en.ID_ENDERECO
                WHERE en.CEP IS NOT NULL 
                AND en.CEP <> ''
                AND (e.LATITUDE IS NULL OR e.LONGITUDE IS NULL)
            """)
            results = cursor.fetchall()
            
            # Construir endereço concatenado no Python
            formatted_results = []
            for row in results:
                empresa_id, logr, num, bairro, cidade, cep = row
                
                # Concatenar endereço no Python
                parts = []
                if logr: parts.append(logr)
                if num: parts.append(num)
                if bairro: parts.append(bairro)
                if cidade: parts.append(cidade)
                
                endereco_concat = ', '.join(parts) if parts else ''
                
                if cep:
                    formatted_results.append((empresa_id, endereco_concat, cep))
            
            return formatted_results
    
    def count_total_search_terms(self) -> int:
        """Conta total de termos no banco"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM TB_TERMOS_BUSCA")
            return cursor.fetchone()[0]
    
    def get_processing_statistics(self) -> Dict[str, int]:
        """Obtém estatísticas completas do processamento"""
        with self._get_connection() as conn:
            cursor = conn.cursor()

            # Termos
            cursor.execute("SELECT COUNT(*) FROM TB_TERMOS_BUSCA")
            total_termos = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM TB_TERMOS_BUSCA WHERE STATUS_PROCESSAMENTO = 'CONCLUIDO'")
            termos_concluidos = cursor.fetchone()[0]

            # Empresas
            cursor.execute("SELECT COUNT(*) FROM TB_EMPRESAS")
            total_empresas = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM TB_EMPRESAS WHERE STATUS_COLETA = 'COLETADO'")
            empresas_coletadas = cursor.fetchone()[0]

            # E-mails e telefones
            cursor.execute("SELECT COUNT(*) FROM TB_EMAILS")
            total_emails = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM TB_TELEFONES")
            total_telefones = cursor.fetchone()[0]

            return {
                'termos_total': total_termos,
                'termos_concluidos': termos_concluidos,
                'termos_pendentes': total_termos - termos_concluidos,
                'empresas_total': total_empresas,
                'empresas_coletadas': empresas_coletadas,
                'emails_total': total_emails,
                'telefones_total': total_telefones,
                'progresso_pct': round((termos_concluidos / total_termos * 100), 1) if total_termos > 0 else 0
            }
    
    # ===== CEP ENRICHMENT =====
    
    def create_cep_enrichment_task(self, empresa_id: int, endereco_id: int):
        """Cria tarefa de enriquecimento CEP"""
        if not endereco_id:
            return
            
        with self._get_connection() as conn:
            cursor = conn.cursor()
            
            # Verificar se já existe tarefa para esta empresa
            cursor.execute("SELECT ID_CEP_ENRICHMENT FROM TB_CEP_ENRICHMENT WHERE ID_EMPRESA = ?", empresa_id)
            existing = cursor.fetchone()
            
            if not existing:
                cursor.execute("""
                               INSERT INTO TB_CEP_ENRICHMENT (ID_EMPRESA, ID_ENDERECO, STATUS_PROCESSAMENTO, TENTATIVAS)
                               VALUES (?, ?, 'PENDENTE', 0)
                               """, empresa_id, endereco_id)
                conn.commit()
    
    def get_pending_cep_enrichment_tasks(self) -> List[Dict[str, Any]]:
        """Obtém tarefas de enriquecimento CEP pendentes"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                           SELECT c.ID_CEP_ENRICHMENT, c.ID_EMPRESA, c.ID_ENDERECO, emp.SITE_URL,
                                  end.LOGRADOURO, end.NUMERO, end.COMPLEMENTO, end.BAIRRO, end.CIDADE, end.ESTADO, end.CEP
                           FROM (TB_CEP_ENRICHMENT c 
                           INNER JOIN TB_EMPRESAS emp ON c.ID_EMPRESA = emp.ID_EMPRESA)
                           INNER JOIN TB_ENDERECOS end ON c.ID_ENDERECO = end.ID_ENDERECO
                           WHERE c.STATUS_PROCESSAMENTO = 'PENDENTE'
                           ORDER BY c.ID_CEP_ENRICHMENT
                           """)
            
            tasks = []
            for row in cursor.fetchall():
                from src.domain.models.address_model import AddressModel
                address = AddressModel(
                    logradouro=row[4] or "",
                    numero=row[5] or "",
                    complemento=row[6] or "",
                    bairro=row[7] or "",
                    cidade=row[8] or "",
                    estado=row[9] or "",
                    cep=row[10] or ""
                )
                
                tasks.append({
                    'id_cep_enrichment': row[0],
                    'id_empresa': row[1], 
                    'id_endereco': row[2],
                    'site_url': row[3],
                    'address_model': address
                })
            
            return tasks
    
    def update_cep_enrichment_success(self, id_cep_enrichment: int):
        """Atualiza sucesso do enriquecimento CEP"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                           UPDATE TB_CEP_ENRICHMENT
                           SET STATUS_PROCESSAMENTO = 'CONCLUIDO', DATA_PROCESSAMENTO = Date(),
                               TENTATIVAS = TENTATIVAS + 1
                           WHERE ID_CEP_ENRICHMENT = ?
                           """, id_cep_enrichment)
            conn.commit()
    
    def update_cep_enrichment_error(self, id_cep_enrichment: int, erro_descricao: str):
        """Atualiza erro no enriquecimento CEP"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                           UPDATE TB_CEP_ENRICHMENT
                           SET STATUS_PROCESSAMENTO = 'ERRO', DATA_PROCESSAMENTO = Date(),
                               TENTATIVAS = TENTATIVAS + 1, ERRO_DESCRICAO = ?
                           WHERE ID_CEP_ENRICHMENT = ?
                           """, erro_descricao, id_cep_enrichment)
            conn.commit()
    
    def get_cep_enrichment_stats(self) -> Dict[str, int]:
        """Obtém estatísticas de enriquecimento CEP"""
        try:
            conn = self._get_connection()
            cursor = conn.cursor()
            
            cursor.execute("SELECT COUNT(*) FROM TB_CEP_ENRICHMENT WHERE STATUS_PROCESSAMENTO = 'CONCLUIDO'")
            concluidos = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM TB_CEP_ENRICHMENT WHERE STATUS_PROCESSAMENTO = 'PENDENTE'")
            pendentes = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM TB_CEP_ENRICHMENT WHERE STATUS_PROCESSAMENTO = 'ERRO'")
            erros = cursor.fetchone()[0]
            
            cursor.close()
            
            total = concluidos + pendentes + erros
            
            return {
                'total': total,
                'concluidos': concluidos,
                'pendentes': pendentes,
                'erros': erros,
                'percentual': round((concluidos / max(total, 1)) * 100, 1)
            }
        except Exception as e:
            print(f"[AVISO] Erro ao obter estatísticas CEP: {e}")
            return {'total': 0, 'concluidos': 0, 'pendentes': 0, 'erros': 0, 'percentual': 0}
    
    # ===== ADDRESS ENRICHMENT =====
    
    def update_endereco_corrected(self, endereco_id: int, corrected_address) -> None:
        """Atualiza endereço corrigido na TB_ENDERECOS"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE TB_ENDERECOS 
                SET LOGRADOURO = ?, NUMERO = ?, COMPLEMENTO = ?, BAIRRO = ?, CIDADE = ?, ESTADO = ?
                WHERE ID_ENDERECO = ?
            """, 
                corrected_address.logradouro,
                corrected_address.numero,
                corrected_address.complemento,
                corrected_address.bairro,
                corrected_address.cidade,
                corrected_address.estado,
                endereco_id
            )
            conn.commit()
    
    def update_endereco_enriched(self, empresa_id: int, enriched_address) -> None:
        """Atualiza endereço enriquecido na TB_ENDERECOS e cria tarefa de geolocalização"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            
            # Obter ID_ENDERECO da empresa
            cursor.execute("SELECT ID_ENDERECO FROM TB_EMPRESAS WHERE ID_EMPRESA = ?", empresa_id)
            result = cursor.fetchone()
            if not result:
                return
            endereco_id = result[0]
            
            # Atualizar TB_ENDERECOS
            cursor.execute("""
                UPDATE TB_ENDERECOS 
                SET LOGRADOURO = ?, NUMERO = ?, COMPLEMENTO = ?, BAIRRO = ?, CIDADE = ?, ESTADO = ?
                WHERE ID_ENDERECO = ?
            """, 
                enriched_address.logradouro,
                enriched_address.numero,
                enriched_address.complemento,
                enriched_address.bairro,
                enriched_address.cidade,
                enriched_address.estado,
                endereco_id
            )
            conn.commit()
            
            # Criar tarefas de enriquecimento CEP e geolocalização
            self.create_cep_enrichment_task(empresa_id, endereco_id)
            self.create_geolocation_task(empresa_id, endereco_id)
    
    def update_empresa_endereco_concatenado(self, empresa_id: int, endereco_completo: str) -> None:
        """Atualiza endereço concatenado na TB_EMPRESAS (REMOVIDO - campo não existe)"""
        # Campo ENDERECO_CONCATENADO não existe na TB_EMPRESAS
        # Pulando atualização para evitar erro SQL
        pass
    
    def fetch_all(self, query: str, params: list = None) -> List[tuple]:
        """Executa SELECT e retorna todas as linhas como tuplas"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            return cursor.fetchall()
    
    def fetch_one(self, query: str, params: list = None) -> tuple:
        """Executa SELECT e retorna uma linha como tupla"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            return cursor.fetchone()
    
    def get_company_collection_statistics(self) -> Dict[str, int]:
        """Obtém estatísticas detalhadas de coleta de empresas"""
        try:
            with self._get_connection() as conn:
                cursor = conn.cursor()
                
                # Empresas visitadas (total de empresas na TB_EMPRESAS)
                cursor.execute("SELECT COUNT(*) FROM TB_EMPRESAS")
                visitadas = cursor.fetchone()[0]
                
                # Empresas coletadas (com dados)
                cursor.execute("SELECT COUNT(*) FROM TB_EMPRESAS WHERE STATUS_COLETA = 'COLETADO'")
                coletadas = cursor.fetchone()[0]
                
                # Empresas não coletadas
                cursor.execute("SELECT COUNT(*) FROM TB_EMPRESAS WHERE STATUS_COLETA = 'NAO_COLETADO'")
                nao_coletadas = cursor.fetchone()[0]
                
                # Taxa de coleta
                taxa_coleta_pct = round((coletadas / max(visitadas, 1)) * 100, 1)
                
                return {
                    'visitadas': visitadas,
                    'coletadas': coletadas,
                    'nao_coletadas': nao_coletadas,
                    'taxa_coleta_pct': taxa_coleta_pct
                }
        except Exception as e:
            print(f"[AVISO] Erro ao obter estatísticas de empresas: {e}")
            return {
                'visitadas': 0,
                'coletadas': 0,
                'nao_coletadas': 0,
                'taxa_coleta_pct': 0
            }
